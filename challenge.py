from openpyxl import load_workbook
from complete_form import charge_form
from complete_email import gmail_send_message

#abro el archivo xlsx, extraigo los datos y luego lo cierro
file = "./Base Seguimiento Observ Auditoría al_30042021.xlsx"
wb = load_workbook(file)
sheet = wb.active
wb.close()

#itero sobre una lista de filas
for row in sheet.iter_rows(min_row=1, max_row=45, max_col=10):
    #evaluo la columna "J"(estado)
    match str.lower(row[9].value):
        case "regularizado":
            #print("subir la información al formulario")
            charge_form(row)
        case "atrasado":
            #print(" enviar un mail ")
            gmail_send_message(row)
        case 'pendiente':
            #en este caso debe ser ignorado
            continue

    print('--- linea ---')



a = input("ingresa algo")